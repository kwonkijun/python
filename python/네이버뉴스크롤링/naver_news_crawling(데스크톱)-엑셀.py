import requests as req
from bs4 import BeautifulSoup
import os
import openpyxl
import pyautogui
from openpyxl.styles import Alignment, Font

keyword = pyautogui.prompt(text='검색어를 입력하세요', title='Message', default='입력하세요')

page_num = 0
# 3 페이지 까지 크롤링
news_title_list = []
news_content_list = []

while (page_num*10 + 1) < 31:
	url = f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={keyword}&sort=0&start={page_num*10 + 1}"	
	# 사이트에 접속해서 html 데이터 가져오기
	response = req.get(url, headers={"User-Agent" : "Mozilla/5.0"})

	# html 가공하기
	html = response.text
	soup = BeautifulSoup(html, "html.parser")

	headers = soup.select("div.info_group")
	print(f"{page_num+1}번째 페이지 크롤링 중입니다..")

	for header in headers:
		link = header.select("a.info")
		# 네이버뉴스 링크가 있는 기사들만 크롤링
		if(len(link) > 1):
			news_url = link[1].attrs['href']
			response = req.get(news_url, headers={"User-Agent" : "Mozilla/5.0"})
			news_html = response.text
			news_soup = BeautifulSoup(news_html, "html.parser")
			try:
				news_title = news_soup.select_one("#articleTitle").string
				news_content = news_soup.select_one("#articleBodyContents").text

				if "// flash 오류를 우회하기 위한 함수 추가" in news_content:
					news_content = news_content.replace("// flash 오류를 우회하기 위한 함수 추가", "")
					news_content = news_content.replace("function _flash_removeCallback() {}", "")

				news_content = news_content.strip() # 양쪽 공백 제거
				news_title_list.append(news_title)
				news_content_list.append(news_content)
				print(f"뉴스제목\n{news_title}\n본문내용\n{news_content}")
			except:
				pass

	page_num += 1


print("뉴스 데이터 크롤링 완료!")

# 엑셀 파일 생성하기
excel_url = r"C:\Users\스타트코딩\Desktop\main\python\python\네이버뉴스크롤링\naver_news.xlsx"
if not os.path.exists(excel_url):
	openpyxl.Workbook().save(excel_url)
book = openpyxl.load_workbook(excel_url)

sheet = book.create_sheet()
sheet.title = keyword
# 열 너비 조절
sheet.column_dimensions["A"].width = 60
sheet.column_dimensions["B"].width = 120

row_num = 1
for i in range(len(news_title_list)):
	sheet.cell(row=row_num, column=1).value = news_title_list[i]
	sheet.cell(row=row_num, column=1).font = Font(size = 16, bold = True)
	sheet.cell(row=row_num, column=2).value = news_content_list[i]
	sheet.cell(row=row_num, column=2).alignment = Alignment(wrap_text = True)
	row_num += 1

book.save(excel_url)
print("엑셀 파일 저장 완료")