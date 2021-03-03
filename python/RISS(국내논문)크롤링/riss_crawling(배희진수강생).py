from selenium import webdriver
import time
import pyautogui
import openpyxl
import os

driver = webdriver.Chrome(r'C:\chromedriver.exe')

keyword = pyautogui.prompt(text="검색어를 입력하세요", title='입력창', default='3글자입력')

url = f"http://www.riss.kr/search/Search.do?query={keyword}"

driver.get(url)

save_path = r'C:\Users\스타트코딩\Desktop\main\python\python\RISS(국내논문)크롤링\논문크롤링결과.xlsx'

# 엑셀 생성 (파일이 없으면 만들고, 있으면 만들지 않는다)
if not os.path.exists(save_path):
    openpyxl.Workbook().save(save_path)

# 엑셀 불러오기
workbook = openpyxl.load_workbook(save_path)

# 시트 생성
sheet = workbook.create_sheet(keyword)
# 국내학술논문 버튼 클릭
driver.find_element_by_css_selector("div.divTabMenu > ul > li.tabM1").click()
time.sleep(1)

# 링크 클릭
titles = driver.find_elements_by_css_selector("div.cont > p.title > a")
# print(titles)
for i in range(len(titles)):
    title = driver.find_elements_by_css_selector("div.cont > p.title > a")
    title[i].click()
    print(f"{i+1}번째 입니다.")
    #  논문이름 발행연도 저자 부가정보
    m_title = driver.find_element_by_css_selector("li > div > p").text
    print('논문이름 : ',m_title)    
    name = driver.find_element_by_css_selector("li > div > p").text
    print('저자 : ',name)
    date = driver.find_elements_by_css_selector("div.infoDetailL > ul >li >div")
    print('발행연도 : ',date[4].text)
    detail = driver.find_elements_by_css_selector("div.innerCont")
    print('부가정보 : ',detail[0].text)


    sheet[f'A{(i+1)}'] = m_title
    sheet[f'B{(i+1)}'] = name
    sheet[f'C{(i+1)}'] = date[4].text
    sheet[f'D{(i+1)}'] = detail[0].text
    workbook.save(save_path)

    time.sleep(1)
    driver.back()