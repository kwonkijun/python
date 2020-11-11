import requests as req
import urllib.request
from bs4 import BeautifulSoup
import os
import openpyxl
import datetime
from openpyxl.drawing.image import Image

# 이미지 저장할 폴더 생성
if not os.path.exists("./멜론이미지"):
	os.mkdir("./멜론이미지")

url = "https://www.melon.com/chart/index.htm"
response = req.get(url, headers={"User-Agent" : "Mozilla/5.0"})

# 406 에러 = 서버에서 비정상적인 접근 판단 
if response.status_code == 200 :
	html = response.text
	soup = BeautifulSoup(html, "html.parser")
	titles = soup.select("div.ellipsis.rank01 > span > a")
	singers = soup.select("span.checkEllipsis")
	albums = soup.select("div.ellipsis.rank03 > a")
	imgs = soup.select("a.image_typeAll > img")
	
	# 엑셀 파일 생성
	if not os.path.exists("./멜론_크롤링.xlsx"):
		openpyxl.Workbook().save("./멜론_크롤링.xlsx")

	book = openpyxl.load_workbook("./멜론_크롤링.xlsx")

	# 쓸데없는 시트 삭제하기
	if "Sheet" in book.sheetnames:
		book.remove(book["Sheet"])

	sheet = book.create_sheet()
	now = datetime.datetime.now()
	sheet.title = f"{now.year}년{now.month}월{now.day}일{now.hour}시{now.minute}분"
	
	# 열 너비 조절
	sheet.column_dimensions["A"].width = 15
	sheet.column_dimensions["B"].width = 50
	sheet.column_dimensions["C"].width = 35
	sheet.column_dimensions["D"].width = 55

	row_num = 1
	for i in range(len(titles)):
		title = titles[i].string
		singer = singers[i].text
		album = albums[i].string
		
		if( '.' in album):
			album = album.replace(".", "")
		if( '?' in album):
			album = album.replace("?", "")

		img_path = f"./멜론이미지/{album}.png"
		urllib.request.urlretrieve(imgs[i].attrs['src'], img_path)
		print(f" {i+1}위. {title} {singer} {album} {imgs[i].attrs['src']}")
		# 엑셀에 크롤링 결과 출력
		img_for_excel = Image(img_path)
		sheet.add_image(img_for_excel, f"A{row_num}")
		sheet.cell(row=row_num, column=2).value = title
		sheet.cell(row=row_num, column=3).value = singer
		sheet.cell(row=row_num, column=4).value = album

		# 행 높이 조절하기
		sheet.row_dimensions[row_num].height = 90
		# 매행 마다 크롤리 결과를 저장한다 (중간에 오류나는 것을 방지)
		book.save("./멜론_크롤링.xlsx")
		row_num += 1