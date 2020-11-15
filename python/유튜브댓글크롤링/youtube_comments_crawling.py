from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
import openpyxl
import os
# 에러 메시지 안뜨게 하기
chrome_opt = webdriver.ChromeOptions()
chrome_opt.add_argument('--disable-gpu')

driver = webdriver.Chrome(r"C:\Users\Administrator\Desktop\기준\코딩교육\자료정리\파이썬\python\selenium\chromedriver.exe", options=chrome_opt)
driver.get("https://www.youtube.com/watch?v=UKCiuyeoYno&ab_channel=%EC%9D%BC%ED%97%A5%ED%83%80%EB%A5%B4TV")
time.sleep(3)
.
# 스크롤 내리기
# driver.find_element_by_css_selector("body").send_keys(Keys.PAGE_DOWN)

# 스크롤 끝까지 내리기
driver.find_element_by_css_selector("body").send_keys(Keys.END)
time.sleep(5)

# 제목 
title = driver.find_element_by_css_selector(".title.style-scope.ytd-video-primary-info-renderer").text

author = driver.find_elements_by_css_selector("#author-text")
content = driver.find_elements_by_css_selector("#content-text")
likes = driver.find_elements_by_css_selector("#vote-count-middle")

# 데이터
author_list = []
content_list = []
likes_list = []
i = 0

while True:
	try:
		author_list.append(author[i].text)
		content_list.append(content[i].text)
		likes_list.append(likes[i].text)
		print(f"{i+1}번째 댓글 {author[i].text} {content[i].text} {likes[i].text}")
	except:
		print("크롤링 완료!!!")		
		excel_url = r"C:\Users\Administrator\Desktop\기준\코딩교육\자료정리\파이썬\python\유튜브댓글크롤링\youtube_result.xlsx"
		# 엑셀 파일 생성하기
		if not os.path.exists(excel_url):
			openpyxl.Workbook().save(excel_url)
		book = openpyxl.load_workbook(excel_url)
		# 쓸데없는 시트 삭제하기
		if "Sheet" in book.sheetnames:
			book.remove(book["Sheet"])
		
		sheet = book.create_sheet()
		sheet.title = title

		row_num = 1
		for i in range(len(author_list)):
			sheet.cell(row=row_num, column=1).value = author_list[i]
			sheet.cell(row=row_num, column=2).value = content_list[i]
			sheet.cell(row=row_num, column=3).value = likes_list[i]
			row_num += 1
		
		book.save(excel_url)
		print("엑셀 파일 저장 완료")
		break
	i += 1
	if(i % 20 == 0):
		driver.find_element_by_css_selector("body").send_keys(Keys.END)
		time.sleep(5)
		author = driver.find_elements_by_css_selector("#author-text")
		content = driver.find_elements_by_css_selector("#content-text")
		likes = driver.find_elements_by_css_selector("#vote-count-middle")

	