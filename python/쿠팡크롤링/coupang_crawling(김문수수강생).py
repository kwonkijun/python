import requests
from bs4 import BeautifulSoup
import pyautogui
import openpyxl
import os

keyword = pyautogui.prompt(text='검색어를 입력하세요', title='Message')
last_page = int(pyautogui.prompt(text='몇 페이지까지 크롤링?', title='Message'))

link_list = []
brand_list = [] 
SKU_list = []
price_list = []

for page in range(1, last_page + 1):
    url = f'https://www.coupang.com/np/search?q={keyword}&page={page}'
    response = requests.get(url, headers ={'User-Agent' : 'Mozilla/5.0'})    
    response.raise_for_status
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    links = soup.select('#productList > li > a')   
    for i in range(len(links)):
        pre_link = links[i].attrs['href']
        link = 'https://www.coupang.com/' + pre_link
        link_list.append(link)
        response = requests.get(link, headers ={'User-Agent' : 'Mozilla/5.0'})    
        response.raise_for_status
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        brand = soup.select('#contents > div.prod-atf > div > div > a')
        SKU = soup.select_one('#contents > div.prod-atf > div > div > div.prod-buy-header > h2').text
        coupon_price = soup.select_one('div.prod-coupon-price.prod-major-price > span > strong')  

        if len(brand) > 0:
            brand = soup.select_one('#contents > div.prod-atf > div > div > a').text
            if len(coupon_price) > 2:
                price = soup.select_one('div.prod-coupon-price.prod-major-price > span > strong').text
                print(brand, SKU, price)
                brand_list.append(brand)
                SKU_list.append(SKU)
                price_list.append(price)
            else:
                try:
                    brand = soup.select_one('div.prod-vendor-container > div > div.prod-sale-vendor > a').text
                    price = soup.select_one('div.prod-sale-price.prod-major-price > span.total-price > strong').text
                    print(brand, SKU, price)
                    brand_list.append(brand)
                    SKU_list.append(SKU)
                    price_list.append(price)
                except AttributeError:
                    brand = '수기입력요망'
                    price = '수기입력요망'
                    print(brand, SKU, price)
                    brand_list.append(brand)
                    SKU_list.append(SKU)
                    price_list.append(price)
        else:
            brand = '브랜드미입력'
            print(brand, SKU, price)
            brand_list.append(brand)
            SKU_list.append(SKU)
            price_list.append(price)

    print(page, '페이지', '크롤링 완료')

save_path = r"C:\Users\스타트코딩\Desktop\main\python\python\쿠팡크롤링\쿠팡크롤러_결과.xlsx"

if not os.path.exists(save_path):
    openpyxl.Workbook().save(save_path)

workbook = openpyxl.load_workbook(save_path)

sheet = workbook.create_sheet(keyword) #Ctrl 누르고 keyword 누르면, 변수 정의내렸던 곳으로 즐겨찾기 된다. 
sheet['A1'] = '링크'
sheet['B1'] = '브랜드'
sheet['C1'] = 'SKU'
sheet['D1'] = '가격'

for i in range(len(link_list)):
    sheet[f'A{i+2}'] = link_list[i]
    sheet[f'B{i+2}'] = brand_list[i]
    sheet[f'C{i+2}'] = SKU_list[i]
    sheet[f'D{i+2}'] = price_list[i]

workbook.save(save_path)

# 이제 결과 엑셀에 옮기자!

# 애자일하게 먼저 큰 틀 잡기
    # 키워드를 검색해 관련 상품 검색결과 페이지로 이동하기
    # 검색 결과 페이지에서, url타고 들어가기 - 타고 들어간 url 크롤링
        # 거기서 브랜드, 상품명, 가격,  추출하기... 
        # 이걸 어떻게 반복문으로 풀어낼지 생각해보자
    # 이후 엑셀로 결과 출력하기

# 디테일까지 보완하기
    # 데이터 없을 땐 오류 처리 뜰 텐데, 이거 어떻게 해결할 거냐?