import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side

path = r"C:\Users\스타트코딩\Desktop\main\python\업무자동화\엑셀자동화\01_광고비_데이터_병합.xlsx"

book = openpyxl.load_workbook(path)
sheet = book.active

font = Font(name="맑은 고딕", size=12, bold=True)
alignment_center = Alignment(horizontal="center")
color_orange = PatternFill(patternType="solid", fgColor=Color("FFD732"))
color_grey = PatternFill(patternType="solid", fgColor=Color("c8c8c8"))
border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
sheet["A14"] = "합계"
sheet["A14"].font = font
sheet["A14"].alignment = alignment_center
sheet["A14"].fill = color_orange
sheet["A14"].border = border

sheet["B14"] = "=SUM(B2:B13)"  # 엑셀 수식 넣기 (그냥 넣어주면된다)
sheet["C14"] = "=SUM(C2:C13)"
sheet["D14"] = "=SUM(D2:D13)"

for row in sheet["B14:D14"]:
    for cell in row:
        cell.font = font
        cell.alignment = alignment_center
        cell.fill = color_orange
        cell.border = border

for row in sheet["B2:D13"]:
    for cell in row:
        cell.font = font
        cell.alignment = alignment_center
        cell.fill = color_grey
        cell.border = border

book.save(path)